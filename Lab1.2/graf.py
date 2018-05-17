from openpyxl import load_workbook
from matplotlib import pyplot
#from openpyxl import *
wb = load_workbook('data_analysis_lab.xlsx')



print(wb.worksheets)
print(wb.sheetnames)
#print(L[0])
sheet = wb['Data']

sheet['A'][1:]
print(sheet['A'][2].value)

def getvalue(x): return x.value


m1 = map(getvalue, sheet['A'][1:])
m2 = map(getvalue, sheet['C'][1:])
m3 = map(getvalue, sheet['D'][1:])
list_x = list(m1)
list_y = list(m2)
list_z = list(m3)
pyplot.plot(list_x, list_y, label="Temp")
pyplot.plot(list_x, list_z, label="Temp2")
pyplot.show()

